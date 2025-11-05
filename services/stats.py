from models import UserQuery, UserStat, DailyStat, SessionLocal
from datetime import datetime, date, timedelta

from typing import Dict, List, Tuple
from utils.logging import logger
from sqlalchemy.sql import func

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def save_user_query(
        user_id: int, username: str, first_name: str,
        last_name: str,
        question: str,
        answer: str,
        section: str = None):
    """Сохраняет запрос пользователя в базу данных"""
    try:
        db = SessionLocal()

        # Сохраняем запрос
        query = UserQuery(
            user_id=user_id,
            username=username,
            first_name=first_name,
            last_name=last_name,
            question=question,
            answer=answer,
            section=section
        )
        db.add(query)

        # Обновляем статистику пользователя
        user_stat = db.query(UserStat).filter(
            UserStat.user_id == user_id
            ).first()
        if user_stat:
            user_stat.total_queries += 1
            user_stat.last_seen = datetime.utcnow()
        else:
            user_stat = UserStat(
                user_id=user_id,
                username=username,
                first_name=first_name,
                last_name=last_name,
                total_queries=1
            )
            db.add(user_stat)

        # Обновляем дневную статистику
        today = date.today()
        daily_stat = db.query(DailyStat).filter(
            DailyStat.stat_date == today
            ).first()
        if daily_stat:
            daily_stat.total_queries += 1
            # Обновляем самое популярное раздел
            if section:
                # Здесь можно добавить логику
                pass
        else:
            # Для нового дня нужно посчитать уникальных пользователей
            unique_users = db.query(UserQuery.user_id).filter(
                func.date(UserQuery.created_at) == today
            ).distinct().count()

            daily_stat = DailyStat(
                stat_date=today,
                total_queries=1,
                unique_users=unique_users
            )
            db.add(daily_stat)

        db.commit()

    except Exception as e:
        logger.error(f"Ошибка сохранения статистики: {e}")
        db.rollback()
    finally:
        db.close()


def get_user_stats(user_id: int = None) -> Dict:
    """Получает статистику по пользователям"""
    try:
        db = SessionLocal()

        if user_id:
            # Статистика конкретного пользователя
            user_stat = db.query(UserStat).filter(
                UserStat.user_id == user_id
                ).first()
            if user_stat:
                return {
                    "user_id": user_stat.user_id,
                    "username": user_stat.username,
                    "first_name": user_stat.first_name,
                    "last_name": user_stat.last_name,
                    "total_queries": user_stat.total_queries,
                    "first_seen": user_stat.first_seen,
                    "last_seen": user_stat.last_seen
                }
            return {}

        # Общая статистика по всем пользователям
        total_users = db.query(UserStat).count()
        total_queries = db.query(func.sum(
            UserStat.total_queries)).scalar() or 0
        active_users = db.query(UserStat).filter(
            UserStat.last_seen >= datetime.utcnow() - timedelta(days=30)
        ).count()

        return {
            "total_users": total_users,
            "total_queries": total_queries,
            "active_users": active_users,
            "avg_queries_per_user": total_queries / total_users if total_users > 0 else 0
        }

    except Exception as e:
        logger.error(f"Ошибка получения статистики пользователей: {e}")
        return {}
    finally:
        db.close()


def get_daily_stats(days: int = 30) -> List[Dict]:
    """Получает статистику за последние N дней"""
    try:
        db = SessionLocal()

        start_date = date.today() - timedelta(days=days)

        stats = db.query(DailyStat).filter(
            DailyStat.stat_date >= start_date
        ).order_by(DailyStat.stat_date.desc()).all()

        result = []
        for stat in stats:
            result.append({
                "date": stat.stat_date,
                "total_queries": stat.total_queries,
                "unique_users": stat.unique_users
            })

        return result

    except Exception as e:
        logger.error(f"Ошибка получения дневной статистики: {e}")
        return []
    finally:
        db.close()


def get_popular_queries(limit: int = 10) -> List[Tuple[str, int]]:
    """Получает самые популярные запросы."""
    try:
        db = SessionLocal()
        popular = db.query(
            UserQuery.question,
            func.count().label('count')
        ).group_by(
            UserQuery.question
            ).order_by(func.count().desc()).limit(limit).all()
        return [(q[0], q[1]) for q in popular]
    except Exception as e:
        logger.error(f"Ошибка получения популярных запросов: {e}")
        return []
    finally:
        db.close()


def get_popular_sections(limit: int = 10) -> List[Tuple[str, int]]:
    """Получает самые популярные разделы/статьи."""
    try:
        db = SessionLocal()
        popular = db.query(
            UserQuery.section,
            func.count().label('count')
        ).filter(
            UserQuery.section.isnot(None),
            UserQuery.section != ''
        ).group_by(UserQuery.section).order_by(func.count().desc()).limit(limit).all()
        return [(section, count) for section, count in popular]
    except Exception as e:
        logger.error(f"Ошибка получения популярных разделов: {e}")
        return []
    finally:
        db.close()


def get_section_stats() -> Dict[str, int]:
    """Получает статистику по разделам"""
    try:
        db = SessionLocal()

        section_stats = db.query(
            UserQuery.section,
            func.count().label('count')
        ).filter(UserQuery.section.isnot(None)).group_by(UserQuery.section).all()

        return {section: count for section, count in section_stats}

    except Exception as e:
        logger.error(f"Ошибка получения статистики разделов: {e}")
        return {}
    finally:
        db.close()


def get_user_queries(user_id: int, period: str) -> List[Dict]:
    """Получает запросы пользователя за выбранный период."""
    db = SessionLocal()
    try:
        now = datetime.utcnow()
        if period == "1_day":
            start_date = now - timedelta(days=1)
        elif period == "1_week":
            start_date = now - timedelta(weeks=1)
        elif period == "1_month":
            start_date = now - timedelta(days=30)
        elif period == "1_year":
            start_date = now - timedelta(days=365)
        else:
            return []

        queries = db.query(UserQuery).filter(
            UserQuery.user_id == user_id,
            UserQuery.created_at >= start_date
        ).order_by(UserQuery.created_at.desc()).all()

        return [
            {
                "question": query.question,
                "answer": query.answer,
                "created_at": query.created_at.strftime("%d.%m.%Y %H:%M"),
            }
            for query in queries
        ]
    finally:
        db.close()


def get_all_queries(period: str) -> List[Dict]:
    """Получает все запросы за выбранный период."""
    db = SessionLocal()
    try:
        now = datetime.utcnow()

        if period == "1_day":
            start_date = now - timedelta(days=1)
            end_date = now
        elif period == "1_week":
            start_date = now - timedelta(weeks=1)
            end_date = now
        elif period == "1_month":
            start_date = now - timedelta(days=30)
            end_date = now
        elif period == "1_year":
            start_date = now - timedelta(days=365)
            end_date = now
        elif len(period.split(".")) == 3:  # дд.мм.гггг
            day, month, year = map(int, period.split("."))
            start_date = datetime(year, month, day)
            end_date = start_date + timedelta(days=1)
        elif len(period.split(".")) == 2:  # мм.гггг
            month, year = map(int, period.split("."))
            start_date = datetime(year, month, 1)
            end_date = datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1)
        elif len(period) == 4:  # гггг
            year = int(period)
            start_date = datetime(year, 1, 1)
            end_date = datetime(year + 1, 1, 1)
        else:
            return []

        queries = db.query(UserQuery).filter(
            UserQuery.created_at >= start_date,
            UserQuery.created_at < end_date
        ).order_by(UserQuery.created_at.desc()).all()

        return [
            {
                "user_id": query.user_id,
                "username": query.username,
                "question": query.question,
                "answer": query.answer,
                "created_at": query.created_at.strftime("%d.%m.%Y %H:%M"),
            }
            for query in queries
        ]
    finally:
        db.close()


def get_all_users() -> List[Dict]:
    """Получает список всех пользователей."""
    db = SessionLocal()
    try:
        users = db.query(UserStat).all()
        return [
            {
                "user_id": user.user_id,
                "username": user.username,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "total_queries": user.total_queries,
            }
            for user in users
        ]
    finally:
        db.close()


def export_queries_to_excel(period: str) -> str:
    """Экспортирует запросы за выбранный период в Excel."""
    try:
        queries = get_all_queries(period)
        if not queries:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Запросы"

        # Заголовки
        headers = ["ID пользователя", "Имя пользователя", "Вопрос", "Ответ", "Дата"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws.column_dimensions[col_letter].width = 20

        # Данные
        for row_num, query in enumerate(queries, 2):
            ws[f"A{row_num}"] = query["user_id"]
            ws[f"B{row_num}"] = query["username"] or "Без имени"
            ws[f"C{row_num}"] = query["question"]
            ws[f"D{row_num}"] = query["answer"]
            ws[f"E{row_num}"] = query["created_at"]

        # Сохранение файла
        filename = f"queries_{period.replace('.', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename

    except Exception as e:
        logger.error(f"Ошибка при экспорте в Excel: {e}")
        return None
